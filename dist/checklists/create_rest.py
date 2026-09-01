import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = '/Users/ruchikandpal/Documents/Default Project/frontend/public/checklists/'

week_fills = {
    1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
}

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def create_wb(name, weeks, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]
    
    headers = ["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"]
    for col, header in enumerate(["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    week_fills = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                  2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
                  3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                  4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row_idx = 2
    for week in weeks:
        for task_idx, task in enumerate(week['tasks']):
            row_data = [week['week'], week['days'], task, week.get('descriptions', [''])[task_idx] if task_idx < len(week.get('descriptions', [])) else '', "Not Started", "", "", "", ""]
            for col_idx, value in enumerate([week['week'], week['days'], task, week.get('descriptions', [''])[task_idx] if task_idx < len(week.get('descriptions', [])) else '', "Not Started", "", "", "", ""], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                week_num = week['week']
                cell.fill = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                             2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
                             3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                             4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}.get(week_num)
            row_idx += 1

    for i, w in enumerate([6, 12, 45, 50, 12, 25, 15, 12, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A2'
    wb.save(f'/Users/ruchikandpal/Documents/Default Project/frontend/public/checklists/{filename}')
    print(f"Created: {filename}")

# PCI-DSS
create_wb('PCI-DSS', [
    {'week': 1, 'days': 'Days 1-7', 'tasks': ['Complete Learning Module 1: PCI-DSS v4.0 Overview & Scoping', 'Define Cardholder Data Environment (CDE) boundaries', 'Map all card data flows (ingress, storage, processing, egress)', 'Identify all system components in scope (servers, apps, network)', 'Document out-of-scope systems with justification', 'Create Card Data Flow Diagram (DFD) for all channels']},
    {'week': 2, 'days': 'Days 8-14', 'tasks': ['Complete Learning Module 2: Network Security & Data Protection', 'Implement firewall/config standards (Req 1): rules, reviews, DMZ', 'Secure system components (Req 2): hardening, vendor defaults, services', 'Protect stored cardholder data (Req 3): encryption, masking, retention', 'Implement key management for encryption (Req 3.5-3.7)', 'Document network diagrams and data flows per Req 1.1-1.2', 'Configure secure protocols (TLS 1.2+) for all card data transmission']},
    {'week': 3, 'days': 'Days 15-21', 'tasks': ['Complete Learning Module 3: Encryption & Vulnerability Management', 'Encrypt transmission over open/public networks (Req 4): TLS everywhere', 'Deploy and maintain anti-malware (Req 5): all systems, updates, logs', 'Develop secure systems/apps (Req 6): SDLC, code review, vuln testing', 'Implement change management for all system components', 'Deploy WAF or equivalent for public-facing web apps (Req 6.4.2)', 'Perform quarterly ASV scans and annual penetration tests (Req 11)']},
    {'week': 4, 'days': 'Days 22-30+', 'tasks': ['Complete Learning Module 4: Access Control, Monitoring & Compliance', 'Implement access control (Req 7-8): least privilege, MFA, unique IDs', 'Restrict physical access (Req 9): facilities, media, visitors', 'Implement logging & monitoring (Req 10): audit trails, SIEM, alerts', 'Test security systems regularly (Req 11): ASV, pen test, IDS/IPS', 'Maintain information security policy (Req 12): roles, awareness, incident response', 'Manage service providers (Req 12.8-12.10): agreements, monitoring, assessments', 'Prepare SAQ/ROC and Attestation of Compliance (AoC)']},
], 'PCI_DSS_v4_Checklist.xlsx')

print("PCI-DSS done")

# SOC2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_wb(name, weeks, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]
    headers = ["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"]
    for col, header in enumerate(["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    week_fills = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                  2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
                  3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                  4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row_idx = 2
    for week in weeks:
        for task_idx, task in enumerate(week['tasks']):
            row_data = [week['week'], week['days'], task, week.get('descriptions', [''])[task_idx] if task_idx < len(week.get('descriptions', [])) else '', "Not Started", "", "", "", ""]
            for col_idx, value in enumerate([week['week'], week['days'], task, week.get('descriptions', [''])[task_idx] if task_idx < len(week.get('descriptions', [])) else '', "Not Started", "", "", "", ""], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                week_num = week['week']
                cell.fill = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                             2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
                             3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                             4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}.get(week_num)
            row_idx += 1
    for i, w in enumerate([6, 12, 45, 50, 12, 25, 15, 12, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A2'
    wb.save(f'/Users/ruchikandpal/Documents/Default Project/frontend/public/checklists/{filename}')
    print(f"Created: {filename}")

# SOC2
create_wb('SOC 2', [
    {'week': 1, 'days': 'Days 1-7', 'tasks': ['Complete Learning Module 1: SOC 2 Fundamentals & Scoping', 'Select applicable TSC: Security (mandatory) + Availability/Confidentiality/PI/Privacy', 'Define system boundary: services, data, infrastructure, people, processes', 'Identify subservice organizations and their roles', 'Perform readiness assessment: gap analysis against TSC', 'Engage CPA firm for readiness assessment or formal audit', 'Define audit period (typically 6-12 months for Type 2)']},
    {'week': 2, 'days': 'Days 8-21', 'tasks': ['CC1: Control Environment - governance, org structure, competence', 'CC2: Communication & Information - policies, commitments, roles', 'CC3: Risk Assessment - risk identification, analysis, mitigation', 'CC4: Monitoring Activities - ongoing eval, separate evaluations', 'CC5: Control Activities - selection, development, technology controls', 'CC6: Logical/Physical Access - auth, authorization, encryption', 'CC7: System Operations - vulnerability mgmt, incident response', 'CC8: Change Management - SDLC, testing, deployment, authorization', 'CC9: Risk Mitigation - business continuity, vendor mgmt']},
    {'week': 3, 'days': 'Days 22-42', 'tasks': ['Availability (A1): capacity planning, environmental protections, DR/BCP', 'Confidentiality (C1): data classification, handling, disposal, NDAs', 'Processing Integrity (PI1): completeness, accuracy, timeliness, authorization', 'Privacy (P1-P8): notice, choice, collection, use, access, disclosure, quality', 'Collect evidence: policies, procedures, logs, configs, tickets, reports', 'Automate evidence collection where possible (scripts, API integrations)', 'Create evidence matrix mapping controls to evidence artifacts']},
    {'week': 4, 'days': 'Days 43-60+', 'tasks': ['Engage CPA firm for SOC 2 Type 2 audit', 'Provide evidence package and management assertion', 'Facilitate CPA testing: inquiry, observation, inspection, re-performance', 'Remediate exceptions/deficiencies during fieldwork', 'Review draft SOC 2 report and management representation letter', 'Obtain final SOC 2 Type 2 report with auditor opinion', 'Distribute report to stakeholders under NDA']},
], 'SOC2_Type2_Checklist.xlsx')

print("SOC2 done")

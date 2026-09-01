import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_wb(name, weeks, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]
    headers = ["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"]
    for col, header in enumerate(["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    week_fills = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                  2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
                  3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                  4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row_idx = 2
    for week in weeks:
        for task_idx, task in enumerate(week['tasks']):
            for col_idx, value in enumerate([week['week'], week['days'], task, week.get('descriptions', [''])[task_idx] if task_idx < len(week.get('descriptions', [])) else '', "Not Started", "", "", "", ""], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                week_num = week['week']
                cell.fill = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                             2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
                             3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                             4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}.get(week_num)
            row_idx += 1
    for i, w in enumerate([6, 12, 45, 50, 12, 25, 15, 12, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A2'
    wb.save(f'/Users/ruchikandpal/Documents/Default Project/frontend/public/checklists/{filename}')
    print(f"Created: {filename}")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# HIPAA
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "HIPAA"
headers = ["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"]
for col, header in enumerate(["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"], 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
week_fills = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
              2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
              3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
              4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
row_idx = 2
for week in [
    {'week': 1, 'days': 'Days 1-7', 'tasks': ['Complete Learning Module 1: HIPAA Fundamentals & Applicability', 'Determine Covered Entity vs Business Associate status', 'Assign Privacy Officer and Security Officer (required)', 'Identify all PHI/ePHI systems, flows, and repositories', 'Conduct comprehensive risk analysis (§164.308(a)(1))', 'Document risk analysis methodology, findings, and risk ratings', 'Identify and document all Business Associates and BAAs needed']},
    {'week': 2, 'days': 'Days 8-14', 'tasks': ['Complete Learning Module 2: Privacy Rule Implementation', 'Develop Notice of Privacy Practices (NPP) and distribution process', 'Implement patient rights: access, amendment, accounting, restrictions', 'Develop minimum necessary policies and workforce training', 'Implement uses/disclosures policies: TPO, authorization, required by law', 'Establish complaint process and breach notification procedures', 'Review and update Business Associate Agreements (BAAs)']},
    {'week': 3, 'days': 'Days 15-28', 'tasks': ['Complete Learning Module 3: Security Rule - Admin & Physical', 'Administrative (§164.308): policies, workforce security, access mgmt', '  - Security awareness training, sanction policy, termination procedures', '  - Information access management, security incident procedures', 'Physical (§164.310): facility access, workstation/device security', '  - Facility access controls, workstation use, device/media controls', 'Develop contingency plan: data backup, DR, emergency mode operations', 'Implement evaluation process: periodic technical/non-technical eval']},
    {'week': 4, 'days': 'Days 22-30+', 'tasks': ['Complete Learning Module 4: Technical Safeguards & Breach Response', 'Technical Safeguards (§164.312): access control, audit controls, integrity', '  - Unique user ID, emergency access, auto logoff, encryption/decryption', '  - Audit controls: audit logs, review process, log integrity', '  - Integrity: PHI integrity mechanisms, encryption at rest/transit', '  - Person/entity authentication, transmission security (TLS, VPN)', 'Breach Notification: risk assessment, 60-day notification, media notice', 'Ongoing: periodic risk analysis, training, policy updates, BAAs review']},
]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HIPAA"
    headers = ["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"]
    for col, header in enumerate(["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    week_fills = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                  2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
                  3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                  4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    row_idx = 2
    for week in [
        {'week': 1, 'days': 'Days 1-7', 'tasks': ['Complete Learning Module 1: HIPAA Fundamentals & Applicability', 'Determine Covered Entity vs Business Associate status', 'Assign Privacy Officer and Security Officer (required)', 'Identify all PHI/ePHI systems, flows, and repositories', 'Conduct comprehensive risk analysis (§164.308(a)(1))', 'Document risk analysis methodology, findings, and risk ratings', 'Identify and document all Business Associates and BAAs needed']},
        {'week': 2, 'days': 'Days 8-14', 'tasks': ['Complete Learning Module 2: Privacy Rule Implementation', 'Develop Notice of Privacy Practices (NPP) and distribution process', 'Implement patient rights: access, amendment, accounting, restrictions', 'Develop minimum necessary policies and workforce training', 'Implement uses/disclosures policies: TPO, authorization, required by law', 'Establish complaint process and breach notification procedures', 'Review and update Business Associate Agreements (BAAs)']},
        {'week': 3, 'days': 'Days 15-28', 'tasks': ['Complete Learning Module 3: Security Rule - Admin & Physical', 'Administrative (§164.308): policies, workforce security, access mgmt', '  - Security awareness training, sanction policy, termination procedures', '  - Information access management, security incident procedures', 'Physical (§164.310): facility access, workstation/device security', '  - Facility access controls, workstation use, device/media controls', 'Develop contingency plan: data backup, DR, emergency mode operations', 'Implement evaluation process: periodic technical/non-technical eval']},
        {'week': 4, 'days': 'Days 22-30+', 'tasks': ['Complete Learning Module 4: Technical Safeguards & Breach Response', 'Technical Safeguards (§164.312): access control, audit controls, integrity', '  - Unique user ID, emergency access, auto logoff, encryption/decryption', '  - Audit controls: audit logs, review process, log integrity', '  - Integrity: PHI integrity mechanisms, encryption at rest/transit', '  - Person/entity authentication, transmission security (TLS, VPN)', 'Breach Notification: risk assessment, 60-day notification, media notice', 'Ongoing: periodic risk analysis, training, policy updates, BAAs review']},
    ]:
        for task_idx, task in enumerate(week['tasks']):
            for col_idx, value in enumerate([week['week'], week['days'], task, week.get('descriptions', [''])[task_idx] if task_idx < len(week.get('descriptions', [])) else '', "Not Started", "", "", "", ""], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                week_num = week['week']
                cell.fill = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                             2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
                             3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                             4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}.get(week_num)
            row_idx += 1
    for i, w in enumerate([6, 12, 45, 50, 12, 25, 15, 12, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = 'A2'
    wb.save('/Users/ruchikandpal/Documents/Default Project/frontend/public/checklists/HIPAA_Checklist.xlsx')
    print("Created: HIPAA_Checklist.xlsx")

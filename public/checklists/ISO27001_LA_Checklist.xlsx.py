import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ISO 27001 LA Checklist"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
week_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Headers
headers = ["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Data
data = [
    # Week 1
    (1, "Days 1-7", "Read ISO 27001:2022 Clauses 4-10 and Annex A controls", "Understand the management system requirements and all 93 controls in Annex A", "Not Started", "Reading notes", "", "", ""),
    (1, "Days 1-7", "Understand ISO 19011 audit principles and ISO 17021 requirements", "Study audit principles, auditor competence, and certification body requirements", "Not Started", "Reading notes", "", "", ""),
    (1, "Days 1-7", "Complete Learning Module 1: ISO 27001 Fundamentals", "Complete ISO 27001 fundamentals training module", "Not Started", "Certificate", "", "", ""),
    (1, "Days 1-7", "Review ISO 27001:2013 vs 2022 changes", "Study Annex SL, risk-based thinking, and new controls", "Not Started", "Comparison document", "", "", ""),
    (1, "Days 1-7", "Study audit terminology", "Learn: criteria, evidence, findings, non-conformities, OFIs", "Not Started", "Glossary", "", "", ""),
    (1, "Days 1-7", "Identify audit types", "Internal, external, surveillance, recertification, combined", "Not Started", "Audit type matrix", "", "", ""),
    
    # Week 2
    (2, "Days 8-14", "Complete Learning Module 2: Audit Planning & Preparation", "Audit planning methodology", "Not Started", "Certificate", "", "", ""),
    (2, "Days 8-14", "Define audit scope, criteria, and objectives", "Define what to audit, against what, and why", "Not Started", "Audit plan draft", "", "", ""),
    (2, "Days 8-14", "Practice creating audit plans and checklists", "Create sample audit plans", "Not Started", "Sample plans", "", "", ""),
    (2, "Days 8-14", "Study document review techniques (Stage 1)", "Stage 1 document review methodology", "Not Started", "Review checklist", "", "", ""),
    (2, "Days 8-14", "Learn risk-based audit approach", "Risk-based sampling and planning", "Not Started", "Risk assessment", "", "", ""),
    (2, "Days 8-14", "Practice writing audit plans", "Create and communicate audit plans", "Not Started", "Sample plans", "", "", ""),
    
    # Week 3
    (3, "Days 15-21", "Complete Learning Module 3: On-site Audit Execution", "On-site audit techniques", "Not Started", "Certificate", "", "", ""),
    (3, "Days 15-21", "Practice opening and closing meetings", "Meeting structure and conduct", "Not Started", "Meeting scripts", "", "", ""),
    (3, "Days 15-21", "Master interviewing techniques", "Effective questioning techniques", "Not Started", "Interview guide", "", "", ""),
    (3, "Days 15-21", "Practice writing audit findings", "Non-conformities, OFIs, positives", "Not Started", "Sample findings", "", "", ""),
    (3, "Days 15-21", "Evaluate audit evidence", "Evidence evaluation against criteria", "Not Started", "Evidence matrix", "", "", ""),
    (3, "Days 15-21", "Daily team meetings and audit log", "Team coordination and logging", "Not Started", "Meeting minutes", "", "", ""),
    
    # Week 4
    (4, "Days 22-30+", "Complete Learning Module 4: Reporting & Certification", "Reporting and certification process", "Not Started", "Certificate", "", "", ""),
    (4, "Days 22-30+", "Practice writing audit reports", "Executive summary, findings, conclusion", "Not Started", "Sample reports", "", "", ""),
    (4, "Days 22-30+", "Evaluate corrective action plans", "Root cause and corrective action evaluation", "Not Started", "CAR templates", "", "", ""),
    (4, "Days 22-30+", "Understand certification decision process", "Certification decision and surveillance", "Not Started", "Process doc", "", "", ""),
    (4, "Days 22-30+", "Practice writing NC reports", "Root cause analysis and CARs", "Not Started", "NC reports", "", "", ""),
    (4, "Days 22-30+", "Understand IRCA/PECB exam requirements", "Exam prep and application", "Not Started", "Exam guide", "", "", ""),
]

# Write data
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        if row_data[0] == 1:
            cell.fill = week_fill
        elif row_data[0] == 2:
            cell.fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
        elif row_data[0] == 3:
            cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        elif row_data[0] == 4:
            cell.fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")

# Column widths
widths = [6, 12, 45, 50, 12, 25, 15, 12, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws.freeze_panes = 'A2'

# Save
wb.save('/Users/ruchikandpal/Documents/Default Project/frontend/public/checklists/ISO27001_LA_Checklist.xlsx')
print("ISO27001_LA_Checklist.xlsx created")

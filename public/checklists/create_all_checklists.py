import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = '/Users/ruchikandpal/Documents/Default Project/frontend/public/checklists/'

# Styles
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
week_fills = {
    1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),      # blue
    2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),      # purple
    3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),      # green
    4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),      # orange
}
header_font_white = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def create_checklist(filename, framework_name, weeks_data, short_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{short_name} Checklist"

    # Headers
    headers = ["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    week_fills = {
        1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
        3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
    }

    headers_list = ["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    week_fills = {
        1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
        3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
    }

    # Write data
    row_idx = 2
    for week in weeks_data:
        for task_idx, (task, desc) in enumerate(week['tasks']):
            row_data = [
                week['week'],
                week['days'],
                task,
                week['task_descriptions'][task_idx] if task_idx < len(week.get('task_descriptions', [])) else '',
                "Not Started",
                "",
                "",
                "",
                ""
            ]
            for col_idx, value in enumerate([week['week'], week['days'], task, week['task_descriptions'][task_idx] if task_idx < len(week.get('task_descriptions', [])) else '', "Not Started", "", "", "", ""], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                week_num = week['week']
                if week_num in week_fills:
                    cell.fill = week_fills[week_num]
            row_idx += 1

    # Column widths
    widths = [6, 12, 45, 50, 12, 25, 15, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    wb.save(os.path.join(OUTPUT_DIR, filename))
    print(f"Created: {filename}")

# ISO 27001 LI
create_checklist('ISO27001_LI_Checklist.xlsx', 'ISO 27001 LI', [
    {'week': 1, 'days': 'Days 1-7', 'title': 'Foundation', 'color': 'indigo',
     'tasks': ['Read ISO 27001:2022 Clauses 4-10 and Annex A controls',
               'Understand PDCA cycle and ISMS implementation methodology',
               'Complete Learning Module 1: ISO 27001 Implementation Fundamentals',
               'Study ISO 27001:2013 vs 2022 changes',
               'Understand ISMS scope definition and organizational context',
               'Identify stakeholders and their information security requirements'],
     'task_descriptions': [
         'Understand the management system requirements and all controls in Annex A',
         'Understand PDCA cycle and ISMS implementation methodology',
         'Complete ISO 27001 implementation fundamentals training module',
         'Study ISO 27001:2013 vs 2022 changes (Annex SL, risk-based thinking)',
         'Understand ISMS scope definition and organizational context',
         'Identify stakeholders and their information security requirements'
     ]},
    {'week': 2, 'title': 'Risk Assessment & Treatment', 'color': 'purple',
     'days': 'Days 8-14',
     'tasks': ['Complete Learning Module 2: Risk Assessment & Treatment',
               'Learn to establish risk assessment methodology and criteria',
               'Practice asset identification, valuation, and risk assessment',
               'Learn to select and justify risk treatment options',
               'Create Statement of Applicability (SoA) with justifications',
               'Define risk acceptance criteria and residual risk acceptance'],
     'task_descriptions': [
         'Learn to establish risk assessment methodology and criteria',
         'Practice asset identification, valuation, and risk assessment',
         'Learn to select and justify risk treatment options',
         'Create Statement of Applicability (SoA) with justifications',
         'Define risk acceptance criteria and residual risk acceptance',
         ''
     ]},
    {'week': 3, 'title': 'ISMS Implementation & Documentation', 'color': 'green',
     'days': 'Days 15-21',
     'tasks': ['Complete Learning Module 3: ISMS Implementation',
               'Develop Information Security Policy and objectives',
               'Implement Annex A controls with procedures and work instructions',
               'Create document control system and record management',
               'Establish training, awareness, and competence programs'],
     'task_descriptions': [
         'Complete ISMS implementation training module',
         'Develop Information Security Policy and objectives',
         'Implement Annex A controls with procedures and work instructions',
         'Create document control system and record management',
         'Establish training, awareness, and competence programs',
         ''
     ]},
    {'week': 4, 'title': 'Internal Audit, Review & Certification', 'color': 'orange',
     'days': 'Days 22-30+',
     'tasks': ['Complete Learning Module 4: Internal Audit & Certification Prep',
               'Plan and conduct internal audit of ISMS',
               'Prepare for and conduct management review meeting',
               'Address non-conformities and implement corrective actions',
               'Prepare for Stage 1 and Stage 2 certification audits'],
     'task_descriptions': [
         'Complete internal audit and certification prep training',
         'Plan and conduct internal audit of ISMS',
         'Prepare for and conduct management review meeting',
         'Address non-conformities and implement corrective actions',
         'Prepare for Stage 1 and Stage 2 certification audits',
         ''
     ]},
], 'ISO27001_LI_Checklist.xlsx', 'ISO 27001 LI')

print("ISO27001_LI_Checklist.xlsx created")

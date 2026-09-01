import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = '/Users/ruchikandpal/Documents/Default Project/frontend/public/checklists/'

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
week_fills = {
    1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
    3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
}
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def make_ws(ws, framework_name, weeks):
    headers = ["Week", "Day Range", "Task", "Description", "Status", "Evidence", "Owner", "Due Date", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row_idx = 2
    for week in weeks:
        for task_idx, task in enumerate(week['tasks']):
            desc = week.get('task_descriptions', [''])[task_idx] if task_idx < len(week.get('task_descriptions', [])) else ''
            row_data = [week['week'], week['days'], task, '', "Not Started", "", "", "", ""]
            for col_idx, value in enumerate([week['week'], week['days'], task, '', "Not Started", "", "", "", ""], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                week_num = week['week']
                if week_num in {1:1,2:2,3:3,4:4}:
                    cell.fill = {1: PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
                                 2: PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
                                 3: PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
                                 4: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")}.get(week_num)
            row_idx += 1

    widths = [6, 12, 45, 50, 12, 25, 15, 12, 30]
    for i, w in enumerate([6, 12, 45, 50, 12, 25, 15, 12, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

def build(name, weeks, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]
    make_ws(ws, name, weeks)
    wb.save(f'/Users/ruchikandpal/Documents/Default Project/frontend/public/checklists/{filename}')
    print(f"Created: {filename}")

# ISO 27001 LI
build('ISO 27001 LI', [
    {'week': 1, 'days': 'Days 1-7', 'tasks': ['Read ISO 27001:2022 Clauses 4-10 and Annex A controls', 'Understand PDCA cycle and ISMS implementation methodology', 'Complete Learning Module 1: ISO 27001 Implementation Fundamentals', 'Study ISO 27001:2013 vs 2022 changes', 'Understand ISMS scope definition and organizational context', 'Identify stakeholders and their information security requirements']},
    {'week': 2, 'days': 'Days 8-14', 'tasks': ['Complete Learning Module 2: Risk Assessment & Treatment', 'Learn to establish risk assessment methodology and criteria', 'Practice asset identification, valuation, and risk assessment', 'Learn to select and justify risk treatment options', 'Create Statement of Applicability (SoA) with justifications', 'Define risk acceptance criteria and residual risk acceptance']},
    {'week': 3, 'days': 'Days 15-21', 'tasks': ['Complete Learning Module 3: ISMS Implementation', 'Develop Information Security Policy and objectives', 'Implement Annex A controls with procedures and work instructions', 'Create document control system and record management', 'Establish training, awareness, and competence programs']},
    {'week': 4, 'days': 'Days 22-30+', 'tasks': ['Complete Learning Module 4: Internal Audit & Certification Prep', 'Plan and conduct internal audit of ISMS', 'Prepare for and conduct management review meeting', 'Address non-conformities and implement corrective actions', 'Prepare for Stage 1 and Stage 2 certification audits']},
], 'ISO27001_LI_Checklist.xlsx')

print("Done creating checklists")
